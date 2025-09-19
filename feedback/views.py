import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Avg
from django.http import HttpResponse
from django.contrib import messages
from django.core.paginator import Paginator

from accounts.views import is_admin
from .filters import LogFilter
from .models import UsageLog, SurveyResponse, SurveyRating
from .forms import FullSurveyForm, SURVEY_QUESTIONS

# ==============================================================================
# Teacher-facing Views
# ==============================================================================

@login_required
def survey_view(request):
    """
    Handles the survey submission. A user can only submit once.
    If they have already submitted, it shows a page informing them.
    An admin can "unlock" their response to allow a re-submission.
    """
    if request.user.role == 'ADMIN':
        return redirect('accounts:admin_dashboard')

    # Check if a locked response already exists for this user
    existing_locked_response = SurveyResponse.objects.filter(user=request.user, is_locked=True).first()
    
    if existing_locked_response and request.method == 'GET':
        # If a locked response exists and it's a GET request, show the "request unlock" page.
        return render(request, 'feedback/survey_request_unlock.html', {'response': existing_locked_response})

    if request.method == 'POST':
        # Find an unlocked (if any) or create a new response instance
        instance_to_update = SurveyResponse.objects.filter(user=request.user, is_locked=False).first()
        form = FullSurveyForm(request.POST, instance=instance_to_update)
        
        if form.is_valid():
            response_instance = form.save(commit=False)
            response_instance.user = request.user
            response_instance.is_locked = True  # Lock the response upon saving
            response_instance.save()
            
            # Clear previous ratings for this response before saving new ones
            SurveyRating.objects.filter(response=response_instance).delete()

            # Save the new individual ratings
            for category, questions in SURVEY_QUESTIONS.items():
                for code, question_text in questions:
                    field_name = f'rating_{code}'
                    rating_value = form.cleaned_data.get(field_name)
                    if rating_value:
                        SurveyRating.objects.create(
                            response=response_instance,
                            question_code=code,
                            rating=int(rating_value)
                        )
            
            messages.success(request, 'ขอบคุณสำหรับความคิดเห็นและการประเมินของท่าน!')
            return redirect('teacher_dashboard')
    else:
        # For a GET request, if no locked response was found, show a blank form
        form = FullSurveyForm()
    
    context = {
        'form': form,
        'survey_questions': SURVEY_QUESTIONS
    }
    return render(request, 'feedback/survey_form_detailed.html', context)

# ==============================================================================
# Admin-facing Views (Reports and Management)
# ==============================================================================

@user_passes_test(is_admin)
def manage_survey_requests(request):
    """
    Admin page to view and manage unlock requests for surveys.
    This lists all responses that are currently "locked".
    """
    locked_responses = SurveyResponse.objects.filter(is_locked=True).select_related('user').order_by('-submitted_at')
    context = {
        'responses': locked_responses
    }
    return render(request, 'admin/survey_manage_requests.html', context)

@user_passes_test(is_admin)
def unlock_survey(request, response_id):
    """
    Action view to unlock a specific survey response, allowing the user to resubmit.
    Requires a POST request for security.
    """
    if request.method == 'POST':
        response_to_unlock = get_object_or_404(SurveyResponse, id=response_id)
        response_to_unlock.is_locked = False
        response_to_unlock.save()
        messages.success(request, f'ปลดล็อกแบบสอบถามสำหรับผู้ใช้ "{response_to_unlock.user.username}" เรียบร้อยแล้ว')
    return redirect('feedback:manage_survey_requests')

# ==============================================================================
# Admin-facing Views (Reports)
# ==============================================================================

@user_passes_test(is_admin)
def survey_results_view(request):
    """
    Displays an aggregated dashboard of the new, detailed survey results.
    """
    responses = SurveyResponse.objects.select_related('user').all().order_by('-submitted_at')
    total_responses = responses.count()
    
    category_stats = {}
    overall_avg = 0
    
    if total_responses > 0:
        question_avg_ratings = SurveyRating.objects.values('question_code').annotate(avg_rating=Avg('rating'))
        question_map = {item['question_code']: {'avg': item['avg_rating']} for item in question_avg_ratings}

        for category, questions in SURVEY_QUESTIONS.items():
            for code, text in questions:
                if code in question_map:
                    question_map[code]['text'] = text

        for category, questions in SURVEY_QUESTIONS.items():
            category_ratings = [question_map[code]['avg'] for code, text in questions if code in question_map and question_map[code].get('avg') is not None]
            if category_ratings:
                category_avg = sum(category_ratings) / len(category_ratings)
                category_stats[category] = {
                    'average': category_avg,
                    'questions': [question_map[code] for code, text in questions if code in question_map]
                }
        
        all_category_averages = [stats['average'] for stats in category_stats.values()]
        if all_category_averages:
            overall_avg = sum(all_category_averages) / len(all_category_averages)

    context = {
        'responses': responses,
        'total_surveys': total_responses,
        'category_stats': category_stats,
        'overall_avg': overall_avg,
        'chart_labels': list(category_stats.keys()),
        'chart_data': [stats['average'] for stats in category_stats.values()],
    }
    return render(request, 'admin/survey_results.html', context)


@user_passes_test(is_admin)
def clear_surveys_view(request):
    """
    Handles the deletion of all survey data (Responses and Ratings).
    """
    if request.method == 'POST':
        SurveyRating.objects.all().delete()
        SurveyResponse.objects.all().delete()
        messages.success(request, 'ข้อมูลแบบสอบถามทั้งหมดถูกลบเรียบร้อยแล้ว')
        return redirect('feedback:survey_results')
    return render(request, 'admin/survey_clear_confirm.html')


@user_passes_test(is_admin)
def export_surveys_excel(request):
    """
    Exports all detailed survey results to an Excel file.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Survey Results'

    columns = ['ผู้ใช้งาน', 'โรงเรียน', 'กลุ่มสาระฯ', 'ระดับการสอน', 'ประสบการณ์', 'เวลาใช้งาน', 'วันที่ส่ง']
    for category, questions in SURVEY_QUESTIONS.items():
        for code, text in questions:
            columns.append(f'คะแนน {code}')
    columns.extend(['สิ่งที่ชื่นชอบ', 'สิ่งที่ควรปรับปรุง', 'ข้อเสนอแนะอื่นๆ'])

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        worksheet.column_dimensions[get_column_letter(col_num)].width = 20

    responses = SurveyResponse.objects.prefetch_related('ratings').select_related('user').all().order_by('-submitted_at')
    for row_num, response in enumerate(responses, 2):
        ratings_map = {rating.question_code: rating.rating for rating in response.ratings.all()}
        
        row_data = [
            response.user.username if response.user else "N/A",
            response.school_name, response.learning_area,
            response.teaching_level, response.teaching_experience,
            response.usage_duration, response.submitted_at.strftime('%Y-%m-%d %H:%M'),
        ]
        
        for category, questions in SURVEY_QUESTIONS.items():
            for code, text in questions:
                row_data.append(ratings_map.get(code, ''))

        row_data.extend([
            response.suggestion_likes, response.suggestion_improvements, response.suggestion_future,
        ])
        
        for col_num, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num).value = cell_value

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="detailed_survey_results.xlsx"'
    workbook.save(response)
    return response

@user_passes_test(is_admin)
def usage_log_view(request):
    """
    Displays a filterable and paginated list of user activities.
    """
    log_list = UsageLog.objects.select_related('user').all().order_by('-action_time')
    log_filter = LogFilter(request.GET, queryset=log_list)
    paginator = Paginator(log_filter.qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'filter': log_filter,
        'page_obj': page_obj,
    }
    return render(request, 'admin/usage_log.html', context)


@user_passes_test(is_admin)
def export_logs_excel(request):
    """
    Exports the filtered usage log data to an Excel file.
    """
    log_list = UsageLog.objects.select_related('user').all().order_by('-action_time')
    log_filter = LogFilter(request.GET, queryset=log_list)
    filtered_logs = log_filter.qs

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Usage Logs'
    columns = ['ผู้ใช้งาน', 'Role', 'กิจกรรม', 'Path', 'IP Address', 'เวลา']
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25

    for row_num, log in enumerate(filtered_logs, 2):
        worksheet.cell(row=row_num, column=1).value = log.user.username if log.user else "N/A"
        worksheet.cell(row=row_num, column=2).value = log.user.get_role_display() if log.user else "N/A"
        worksheet.cell(row=row_num, column=3).value = log.action
        worksheet.cell(row=row_num, column=4).value = log.path
        worksheet.cell(row=row_num, column=5).value = log.ip_address or "-"
        worksheet.cell(row=row_num, column=6).value = log.action_time.strftime('%Y-%m-%d %H:%M:%S')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="usage_logs.xlsx"'
    workbook.save(response)
    return response